import openpyxl
import os
from datetime import datetime

class XlsxExcel():
    def __init__(self, filepath, sheetIndex):
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath, data_only=True)
        self.sheet = self.wb.worksheets[sheetIndex]

    def getCols(self, colIndex, startRowIndex):
        datas = []
        for row in list(self.sheet.rows)[startRowIndex:]:
            datas.append(list(row)[colIndex].value)
        return datas

    def writeCols(self, colIndex, startRowIndex, values):
        for i, value in enumerate(values):
            self.sheet.cell(i + startRowIndex + 1, colIndex + 1).value = value

    def writeRows(self, colIndex, startRowIndex, values):
        for i, value in enumerate(values):
            self.sheet.cell(colIndex + 1, i + startRowIndex + 1).value = value

    def save(self):
        self.wb.save(self.filepath)


project_dir = os.path.dirname(os.path.abspath(__file__))
excel_folder_name = '.doc'
doc_dir = os.path.join(project_dir, excel_folder_name)

if not os.path.exists(doc_dir):
    os.makedirs(doc_dir)
now = datetime.now()
file_created_time = now.strftime('%Y-%m-%d_%H%M')
file_name = f"import_result_file_{file_created_time}.xlsx"
file_path = os.path.join(doc_dir, file_name)

# create a new Excel file
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = '獎金獵人'
workbook.save(file_path)

# add data to worksheet
import_result_file = XlsxExcel(file_path, 0)

column_titles = ['單位名稱', 'Email', '電話', '活動名稱', '活動日期', '網站']

# import_result_file.writeCols(0, 0, ["Fruits"] + test_data_fruits)
import_result_file.writeRows(0, 0, column_titles)


import_result_file.save()


